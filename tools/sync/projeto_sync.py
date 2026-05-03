#!/usr/bin/env python3
"""
projeto_sync.py — Sync incremental do estado do projeto a partir do ClickUp.

Opção D (SessionStart hook): roda ao abrir Claude Code em Claude-NTICS-Projetos, puxa
task-mãe + subtasks + comentários do ClickUp, diffa contra cache, e só chama
Haiku (LLM) quando há delta real. Em dia sem mudança, custo = 0 tokens LLM
(só 1-2 chamadas HTTP ClickUp).

Uso:
    python tools/sync/projeto_sync.py 132-estacao-samarco
    python tools/sync/projeto_sync.py 132-estacao-samarco --quiet
    python tools/sync/projeto_sync.py 132-estacao-samarco --dry-run

Aceita também aliases curtos como '132-samarco' (resolvido por prefixo numérico).
"""

from __future__ import annotations

import argparse
import hashlib
import json
import os
import sys
from datetime import datetime, timezone
from pathlib import Path

import requests
import yaml
import io
from dotenv import load_dotenv

# Windows console defaults to cp1252; forçar UTF-8 para acentos e emojis.
if hasattr(sys.stdout, "reconfigure"):
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except Exception:
        pass

ROOT = Path(__file__).resolve().parents[2]
PROJECTS_DIR = ROOT / "SecondBrain" / "projetos"
AUTOMACOES_ROOT = ROOT.parent / "AUTOMAÇÕES"
AUTOMACOES_ENV = AUTOMACOES_ROOT / ".env"

# Piggyback no auth Google de AUTOMAÇÕES (token.json cacheado).
sys.path.insert(0, str(AUTOMACOES_ROOT / "tools"))

CLICKUP_BASE = "https://api.clickup.com/api/v2"
HAIKU_MODEL = "claude-haiku-4-5-20251001"
ANTHROPIC_API = "https://api.anthropic.com/v1/messages"


def log(msg: str, quiet: bool = False) -> None:
    if not quiet:
        print(msg, file=sys.stderr, flush=True)


def load_env() -> None:
    if AUTOMACOES_ENV.exists():
        load_dotenv(AUTOMACOES_ENV)
    env_local = ROOT / ".env"
    if env_local.exists():
        load_dotenv(env_local, override=True)


def clickup_headers() -> dict:
    token = os.environ.get("CLICKUP_API_KEY") or os.environ.get("CLICKUP_API_TOKEN")
    if not token:
        raise RuntimeError("CLICKUP_API_KEY não encontrado no .env")
    return {"Authorization": token, "Content-Type": "application/json"}


def fetch_task(task_id: str) -> dict | None:
    r = requests.get(
        f"{CLICKUP_BASE}/task/{task_id}",
        headers=clickup_headers(),
        params={"include_subtasks": "true"},
        timeout=30,
    )
    if not r.ok:
        log(f"[warn] fetch_task({task_id}): HTTP {r.status_code} {r.text[:200]}")
        return None
    return r.json()


def fetch_comments(task_id: str) -> list[dict]:
    r = requests.get(
        f"{CLICKUP_BASE}/task/{task_id}/comment",
        headers=clickup_headers(),
        timeout=30,
    )
    if not r.ok:
        return []
    return r.json().get("comments", [])


def snapshot_task(task: dict, comments: list[dict]) -> dict:
    """Reduz payload ClickUp ao que importa para diff + LLM."""
    return {
        "id": task.get("id"),
        "name": task.get("name"),
        "status": (task.get("status") or {}).get("status"),
        "date_updated": task.get("date_updated"),
        "assignees": [a.get("username") for a in task.get("assignees", [])],
        "url": task.get("url"),
        "comments": [
            {
                "id": c.get("id"),
                "author": (c.get("user") or {}).get("username"),
                "date": c.get("date"),
                "text": (c.get("comment_text") or "")[:500],
            }
            for c in comments[-10:]
        ],
    }


def hash_snapshot(obj: dict | list) -> str:
    blob = json.dumps(obj, sort_keys=True, ensure_ascii=False)
    return hashlib.sha256(blob.encode("utf-8")).hexdigest()[:16]


def diff_snapshots(old: dict, new: dict) -> dict:
    """Retorna apenas os campos alterados + novos comentários."""
    delta = {}
    for key in ("status", "name", "assignees"):
        if old.get(key) != new.get(key):
            delta[key] = {"de": old.get(key), "para": new.get(key)}

    old_comment_ids = {c["id"] for c in old.get("comments", [])}
    novos = [c for c in new.get("comments", []) if c["id"] not in old_comment_ids]
    if novos:
        delta["comentarios_novos"] = novos

    return delta


def call_haiku(prompt: str, max_tokens: int = 1500) -> str:
    api_key = os.environ.get("ANTHROPIC_API_KEY")
    if not api_key:
        raise RuntimeError("ANTHROPIC_API_KEY não encontrado no .env")
    r = requests.post(
        ANTHROPIC_API,
        headers={
            "x-api-key": api_key,
            "anthropic-version": "2023-06-01",
            "content-type": "application/json",
        },
        json={
            "model": HAIKU_MODEL,
            "max_tokens": max_tokens,
            "messages": [{"role": "user", "content": prompt}],
        },
        timeout=60,
    )
    if not r.ok:
        raise RuntimeError(f"Haiku API error {r.status_code}: {r.text[:300]}")
    data = r.json()
    return data["content"][0]["text"]


def classify_deltas(project_slug: str, deltas: list[dict], state_context: str) -> dict:
    """Pede ao Haiku para classificar relevância e extrair decisões."""
    prompt = f"""Você é um assistente de coordenação de projetos NTICS.

Projeto: {project_slug}
Contexto atual (state.yaml resumido):
{state_context[:2000]}

Deltas detectadas desde o último sync (ClickUp e/ou Gmail — a chave "fonte" indica "gmail", ausência significa ClickUp):
{json.dumps(deltas, ensure_ascii=False, indent=2)}

Para cada delta, responda em JSON estrito (sem markdown, sem comentário):
{{
  "eventos": [
    {{
      "task_id": "...",
      "tipo": "status_change" | "new_comment" | "name_change" | "assignee_change" | "email_novo" | "email_resposta",
      "fonte": "clickup" | "gmail",
      "categoria": "operacional" | "marca_escopo" | "decisao" | "ruido",
      "resumo": "frase curta do que mudou",
      "deliverable_id_afetado": "A1" ou null (se o nome da task bater com um deliverable conhecido do state.yaml, ex: A1, B2, C1),
      "acao_sugerida": "frase curta do que o coordenador deve fazer OU null",
      "relevancia": "alta" | "media" | "baixa"
    }}
  ],
  "decisoes_capturadas": [
    "frase que parece ser uma decisão formal em comentário (ou vazio)"
  ]
}}

Regras:
- Comentário "ok", "vi", emoji sozinho = baixa relevância, categoria=ruido.
- Mudança de status, menção a aprovação/rejeição/deadline = alta, categoria=operacional.
- Se comentário/email contém "aprovado", "rejeitado", "ajustar", "enviar até", "decidimos", "definido" → categoria=decisao + extrair como decisão.
- Se email/comentário traz informação sobre marca, posicionamento, identidade visual, manual de marca, KV, paleta, tom de voz, escopo do projeto, novos territórios/cidades, mudança de público-alvo → categoria=marca_escopo.
- Tudo o mais que é movimentação de tarefa/entrega no dia a dia → categoria=operacional.
- Nunca invente deliverable_id_afetado. Só preencha se o nome/id da task bater claramente."""

    # 250 tokens por delta como buffer; minimo 2000, maximo 6000
    max_tokens = max(2000, min(6000, 250 * max(1, len(deltas))))
    raw = call_haiku(prompt, max_tokens=max_tokens)
    clean = raw.strip()
    # Tira cerca ```json ... ```
    if clean.startswith("```"):
        clean = clean.strip("`")
        if clean.startswith("json"):
            clean = clean[4:]
        clean = clean.strip()
        if clean.endswith("```"):
            clean = clean[:-3].strip()
    try:
        return json.loads(clean)
    except json.JSONDecodeError:
        # Última tentativa: extrai do primeiro { ao último }
        i, j = clean.find("{"), clean.rfind("}")
        if i != -1 and j != -1 and j > i:
            try:
                return json.loads(clean[i : j + 1])
            except json.JSONDecodeError:
                pass
        raise


def fetch_gmail_threads(query: str, max_results: int = 30) -> list[dict]:
    """Puxa threads Gmail com query (ex: 'from:samarco.com newer_than:2d').
    Retorna lista com id, subject, from, snippet, date. Sem corpo completo (economia)."""
    try:
        from gws.gws_auth import get_credentials  # type: ignore
        from googleapiclient.discovery import build  # type: ignore
    except ImportError as e:
        log(f"[warn] Gmail skip: {e}")
        return []

    try:
        creds = get_credentials()
        service = build("gmail", "v1", credentials=creds, cache_discovery=False)
        resp = service.users().threads().list(
            userId="me", q=query, maxResults=max_results
        ).execute()
    except Exception as e:
        log(f"[warn] Gmail API falhou: {e}")
        return []

    threads = []
    for t in resp.get("threads", []):
        tid = t.get("id")
        try:
            detail = service.users().threads().get(
                userId="me", id=tid, format="metadata",
                metadataHeaders=["Subject", "From", "To", "Date"],
            ).execute()
        except Exception:
            continue
        msgs = detail.get("messages", [])
        if not msgs:
            continue
        last = msgs[-1]
        headers = {h["name"]: h["value"] for h in last.get("payload", {}).get("headers", [])}
        threads.append({
            "id": tid,
            "last_msg_id": last.get("id"),
            "subject": headers.get("Subject", ""),
            "from": headers.get("From", ""),
            "to": headers.get("To", ""),
            "date": headers.get("Date", ""),
            "snippet": (last.get("snippet") or "")[:300],
            "n_msgs": len(msgs),
        })
    return threads


def fetch_planilha_xlsx(file_id: str) -> dict:
    """Baixa .xlsx do Drive e retorna snapshot por aba (qtd de celulas nao-vazias + hash)."""
    try:
        from gws.gws_auth import get_credentials  # type: ignore
        from googleapiclient.discovery import build  # type: ignore
        from googleapiclient.http import MediaIoBaseDownload  # type: ignore
        from openpyxl import load_workbook  # type: ignore
    except ImportError as e:
        log(f"[warn] Planilha skip: {e}")
        return {}

    try:
        creds = get_credentials()
        service = build("drive", "v3", credentials=creds, cache_discovery=False)
        request = service.files().get_media(fileId=file_id)
        buf = io.BytesIO()
        downloader = MediaIoBaseDownload(buf, request)
        done = False
        while not done:
            _, done = downloader.next_chunk()
        buf.seek(0)
        wb = load_workbook(buf, data_only=True)
    except Exception as e:
        log(f"[warn] Planilha fetch falhou: {e}")
        return {}

    snapshot = {"file_id": file_id, "abas": {}}
    for name in wb.sheetnames:
        ws = wb[name]
        rows_nonempty = []
        for row in ws.iter_rows(values_only=True):
            clean = tuple("" if v is None else str(v) for v in row)
            if any(c.strip() for c in clean):
                rows_nonempty.append(clean)
        # hash do conteudo preenchido + qtd de linhas preenchidas
        content_blob = "\n".join(["|".join(r) for r in rows_nonempty])
        snapshot["abas"][name] = {
            "linhas_preenchidas": len(rows_nonempty),
            "hash": hashlib.sha256(content_blob.encode("utf-8")).hexdigest()[:16],
            # primeiras 30 linhas preenchidas (para Haiku classificar delta)
            "preview": rows_nonempty[:30],
        }
    return snapshot


def diff_planilha(old: dict, new: dict) -> list[dict]:
    """Retorna abas que mudaram de hash + delta de linhas preenchidas."""
    deltas = []
    old_abas = (old or {}).get("abas", {})
    new_abas = new.get("abas", {})
    for name, new_info in new_abas.items():
        old_info = old_abas.get(name)
        if not old_info:
            deltas.append({
                "fonte": "planilha",
                "aba": name,
                "tipo": "aba_nova",
                "linhas_preenchidas": new_info["linhas_preenchidas"],
                "preview": new_info["preview"][:10],
            })
        elif old_info["hash"] != new_info["hash"]:
            deltas.append({
                "fonte": "planilha",
                "aba": name,
                "tipo": "aba_alterada",
                "linhas_antes": old_info["linhas_preenchidas"],
                "linhas_agora": new_info["linhas_preenchidas"],
                "preview": new_info["preview"][:10],
            })
    return deltas


def diff_gmail(old: list[dict], new: list[dict]) -> list[dict]:
    """Retorna threads com last_msg_id diferente (ou threads totalmente novas)."""
    old_by_id = {t["id"]: t for t in old}
    deltas = []
    for t in new:
        prev = old_by_id.get(t["id"])
        if not prev:
            deltas.append({"tipo": "thread_nova", **t})
        elif prev.get("last_msg_id") != t.get("last_msg_id"):
            deltas.append({
                "tipo": "mensagem_nova_em_thread",
                "msgs_antes": prev.get("n_msgs"),
                "msgs_agora": t.get("n_msgs"),
                **t,
            })
    return deltas


def load_state(project_dir: Path) -> dict:
    return yaml.safe_load((project_dir / "state.yaml").read_text(encoding="utf-8"))


def load_tasks_index(project_dir: Path) -> dict:
    return yaml.safe_load((project_dir / "tasks.yaml").read_text(encoding="utf-8"))


def append_events_log(project_dir: Path, entries: list[str]) -> None:
    log_path = project_dir / "comms" / "events.log"
    log_path.parent.mkdir(exist_ok=True)
    with log_path.open("a", encoding="utf-8") as f:
        for e in entries:
            f.write(e + "\n")


def find_sb_dir(slug: str) -> Path:
    """Resolve pasta canônica em SecondBrain/projetos/ a partir de slug livre.

    Tenta match exato; se falhar, usa prefixo numérico (ex: '132-samarco' → '132-estacao-samarco').
    Cria pasta com slug cru se nada bater (caminho de bootstrap).
    """
    exact = PROJECTS_DIR / slug
    if exact.exists():
        return exact
    prefix = slug.split("-", 1)[0]
    if prefix.isdigit() and PROJECTS_DIR.exists():
        for child in PROJECTS_DIR.iterdir():
            if child.is_dir() and child.name.startswith(f"{prefix}-"):
                return child
    exact.mkdir(parents=True, exist_ok=True)
    return exact


def append_decisoes_sb(slug: str, decisoes: list[str], fonte: str = "sync") -> None:
    """Append decisões capturadas no SecondBrain do projeto."""
    if not decisoes:
        return
    sb_dir = find_sb_dir(slug)
    path = sb_dir / "decisoes.md"
    ts = datetime.now(timezone.utc).isoformat(timespec="minutes")
    with path.open("a", encoding="utf-8") as f:
        for d in decisoes:
            if d and d.strip():
                f.write(f"\n[{ts}] [AUTO] {d.strip()}\n")


def append_execucao_sb(slug: str, eventos_op: list[dict]) -> None:
    """Append eventos operacionais (status, comentários, emails de movimentação) no execucao.md."""
    if not eventos_op:
        return
    sb_dir = find_sb_dir(slug)
    path = sb_dir / "execucao.md"
    ts = datetime.now(timezone.utc).isoformat(timespec="minutes")
    if not path.exists():
        path.write_text(f"# Log de execução — {slug}\n\nAppend-only. Auto-gerado por projeto_sync.py.\n\n", encoding="utf-8")
    with path.open("a", encoding="utf-8") as f:
        for ev in eventos_op:
            tipo = ev.get("tipo", "evento")
            fonte = ev.get("fonte", "?")
            resumo = (ev.get("resumo") or "").strip()
            deliverable = ev.get("deliverable_id_afetado")
            acao = ev.get("acao_sugerida")
            line = f"[{ts}] [auto:{fonte}/{tipo}] {resumo}"
            if deliverable:
                line += f" (deliverable={deliverable})"
            if acao:
                line += f" → {acao}"
            f.write(line + "\n")


def append_brand_block(project_dir: Path, eventos_brand: list[dict]) -> None:
    """Append eventos de marca/escopo num bloco auto-gerado no fim do CLAUDE.md do projeto."""
    if not eventos_brand:
        return
    claude_md = project_dir / "CLAUDE.md"
    if not claude_md.exists():
        return  # não cria do zero — projeto sem CLAUDE.md inicializado é problema separado
    ts = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    marker = "## Atualizações automáticas (auto-gerado)"
    body = claude_md.read_text(encoding="utf-8")
    bloco_novo = []
    for ev in eventos_brand:
        fonte = ev.get("fonte", "?")
        resumo = (ev.get("resumo") or "").strip()
        if resumo:
            bloco_novo.append(f"- **[{ts}] [{fonte}]** {resumo}")
    if not bloco_novo:
        return
    if marker in body:
        # append no fim do arquivo (novo bloco datado)
        novo = body.rstrip() + "\n\n" + "\n".join(bloco_novo) + "\n"
    else:
        novo = body.rstrip() + f"\n\n{marker}\n\n" + "\n".join(bloco_novo) + "\n"
    claude_md.write_text(novo, encoding="utf-8")


def write_cache(project_dir: Path, snapshot: dict) -> None:
    cache_path = project_dir / "_cache" / "clickup-snapshot.json"
    cache_path.parent.mkdir(exist_ok=True)
    cache_path.write_text(
        json.dumps(snapshot, ensure_ascii=False, indent=2), encoding="utf-8"
    )


def read_cache(project_dir: Path) -> dict:
    cache_path = project_dir / "_cache" / "clickup-snapshot.json"
    if not cache_path.exists():
        return {}
    try:
        return json.loads(cache_path.read_text(encoding="utf-8"))
    except json.JSONDecodeError:
        return {}


def sync_project(slug: str, dry_run: bool = False, quiet: bool = False) -> dict:
    project_dir = find_sb_dir(slug)
    if not project_dir.exists():
        raise SystemExit(f"Projeto não encontrado: {project_dir}")

    state = load_state(project_dir)
    tasks_idx = load_tasks_index(project_dir)
    task_mae_id = tasks_idx.get("task_mae", {}).get("id")
    if not task_mae_id:
        raise SystemExit(f"tasks.yaml sem task_mae.id em {slug}")

    log(f"> Sync {slug} — task-mãe {task_mae_id}", quiet)

    # 1. Fetch ClickUp
    task_mae = fetch_task(task_mae_id)
    if not task_mae:
        return {"status": "erro", "motivo": "task-mãe inacessível"}

    comments_mae = fetch_comments(task_mae_id)
    snap_atual = {
        "timestamp": datetime.now(timezone.utc).isoformat(),
        "task_mae": snapshot_task(task_mae, comments_mae),
        "subtasks": [],
        "gmail_threads": [],
    }

    for sub in task_mae.get("subtasks", []):
        sub_id = sub.get("id")
        sub_comments = fetch_comments(sub_id)
        snap_atual["subtasks"].append(snapshot_task(sub, sub_comments))

    # 1b. Fetch Gmail (se houver query configurada no tasks.yaml do projeto)
    gmail_query = tasks_idx.get("gmail_query")
    if gmail_query:
        snap_atual["gmail_threads"] = fetch_gmail_threads(gmail_query)

    # 1c. Fetch planilhas operacionais (lista de file_ids Drive)
    snap_atual["planilhas"] = {}
    for p in tasks_idx.get("planilhas_monitoradas", []) or []:
        fid = p.get("file_id")
        label = p.get("nome", fid)
        if not fid:
            continue
        snap = fetch_planilha_xlsx(fid)
        if snap:
            snap_atual["planilhas"][label] = snap

    # 2. Diff contra cache
    cache = read_cache(project_dir)
    cache_sub_by_id = {s["id"]: s for s in cache.get("subtasks", [])}

    deltas = []
    if cache:
        d_mae = diff_snapshots(cache.get("task_mae", {}), snap_atual["task_mae"])
        if d_mae:
            deltas.append({"task": snap_atual["task_mae"]["name"], "id": snap_atual["task_mae"]["id"], **d_mae})
        for sub in snap_atual["subtasks"]:
            old_sub = cache_sub_by_id.get(sub["id"], {})
            d = diff_snapshots(old_sub, sub)
            if d:
                deltas.append({"task": sub["name"], "id": sub["id"], **d})
        # Gmail deltas
        gmail_deltas = diff_gmail(cache.get("gmail_threads", []), snap_atual["gmail_threads"])
        for gd in gmail_deltas:
            deltas.append({"fonte": "gmail", **gd})
        # Planilha deltas
        old_planilhas = cache.get("planilhas", {}) or {}
        for label, new_snap in snap_atual["planilhas"].items():
            pd = diff_planilha(old_planilhas.get(label, {}), new_snap)
            for d in pd:
                deltas.append({"planilha": label, **d})

    # 3. Sem mudança? Grava cache e sai. Zero custo LLM.
    if not deltas:
        if not dry_run:
            write_cache(project_dir, snap_atual)
        log(f"[ok] Sem mudanças desde último sync. 0 tokens LLM.", quiet)
        return {"status": "sem_mudancas", "deltas": 0}

    log(f"> {len(deltas)} delta(s) detectada(s). Classificando via Haiku...", quiet)

    # 4. Classificar deltas com Haiku
    state_resumo = yaml.safe_dump(
        {
            "fase": state.get("fase"),
            "proxima_acao": state.get("proxima_acao"),
            "blockers": state.get("blockers", []),
            "deliverables": [
                {"id": d.get("id"), "nome": d.get("nome"), "status": d.get("status")}
                for d in state.get("deliverables_design", [])
            ],
        },
        allow_unicode=True,
        sort_keys=False,
    )

    try:
        classificacao = classify_deltas(slug, deltas, state_resumo)
    except Exception as e:
        log(f"[warn] Haiku falhou: {e}. Registrando deltas crus no log.", quiet)
        eventos_cru = []
        for d in deltas:
            label = d.get("task") or d.get("planilha") or d.get("subject") or d.get("fonte") or "delta"
            campos = [k for k in d.keys() if k not in ("task", "id", "fonte", "planilha")]
            eventos_cru.append({
                "task_id": d.get("id"),
                "tipo": d.get("tipo") or "delta_cru",
                "resumo": f"{label} — {campos}",
                "deliverable_id_afetado": None,
                "acao_sugerida": None,
                "relevancia": "media",
            })
        classificacao = {"eventos": eventos_cru, "decisoes_capturadas": []}

    # 5. Append events.log
    ts = datetime.now(timezone.utc).isoformat(timespec="seconds")
    linhas = []
    for ev in classificacao.get("eventos", []):
        linhas.append(
            f"[{ts}] [{ev.get('relevancia','?')}] task={ev.get('task_id')} "
            f"tipo={ev.get('tipo')} — {ev.get('resumo')} "
            f"(deliverable={ev.get('deliverable_id_afetado')}) "
            f"→ {ev.get('acao_sugerida') or '—'}"
        )
    for dec in classificacao.get("decisoes_capturadas", []):
        if dec and dec.strip():
            linhas.append(f"[{ts}] [DECISAO] {dec}")

    if not dry_run:
        append_events_log(project_dir, linhas)
        append_decisoes_sb(slug, classificacao.get("decisoes_capturadas", []))
        # Roteamento por categoria
        eventos = classificacao.get("eventos", [])
        eventos_op = [e for e in eventos if e.get("categoria") == "operacional" and e.get("relevancia") != "baixa"]
        eventos_brand = [e for e in eventos if e.get("categoria") == "marca_escopo"]
        append_execucao_sb(slug, eventos_op)
        append_brand_block(project_dir, eventos_brand)
        write_cache(project_dir, snap_atual)

    log(f"[ok] Sync completo. {len(classificacao.get('eventos', []))} evento(s) registrado(s).", quiet)

    return {
        "status": "sincronizado",
        "deltas": len(deltas),
        "eventos": classificacao.get("eventos", []),
        "decisoes": classificacao.get("decisoes_capturadas", []),
    }


def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("slug", help="Slug do projeto (ex: 132-estacao-samarco; aceita alias '132-samarco')")
    parser.add_argument("--dry-run", action="store_true", help="Não grava cache nem log")
    parser.add_argument("--quiet", action="store_true", help="Silencia logs no stderr")
    parser.add_argument("--json", action="store_true", help="Imprime resultado como JSON no stdout")
    args = parser.parse_args()

    load_env()
    result = sync_project(args.slug, dry_run=args.dry_run, quiet=args.quiet)

    if args.json:
        print(json.dumps(result, ensure_ascii=False, indent=2))
    else:
        if result["status"] == "sincronizado":
            for ev in result.get("eventos", []):
                marker = {"alta": "[!]", "media": "[·]", "baixa": "[ ]"}.get(ev.get("relevancia"), "[?]")
                acao = ev.get("acao_sugerida") or "-"
                try:
                    print(f"{marker} {ev.get('resumo')} -> {acao}")
                except UnicodeEncodeError:
                    print(f"{marker} {ev.get('resumo','').encode('ascii','replace').decode()} -> {acao.encode('ascii','replace').decode()}")


if __name__ == "__main__":
    main()
