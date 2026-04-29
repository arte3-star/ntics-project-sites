#!/usr/bin/env python3
"""
Baixa um arquivo Excel do Google Drive e imprime estrutura + dados.

Uso:
    python tools/sync/read_drive_xlsx.py <file_id> [--sheet <nome>] [--rows N]
"""
import argparse
import io
import sys
from pathlib import Path

ROOT = Path(__file__).resolve().parents[2]
AUTOMACOES = ROOT.parent / "AUTOMAÇÕES"
sys.path.insert(0, str(AUTOMACOES / "tools"))

from gws.gws_auth import get_credentials  # type: ignore
from googleapiclient.discovery import build  # type: ignore
from googleapiclient.http import MediaIoBaseDownload  # type: ignore

try:
    from openpyxl import load_workbook  # type: ignore
except ImportError:
    print("ERRO: pip install openpyxl", file=sys.stderr)
    sys.exit(1)


def download_xlsx(file_id: str) -> bytes:
    creds = get_credentials()
    service = build("drive", "v3", credentials=creds, cache_discovery=False)
    request = service.files().get_media(fileId=file_id)
    buf = io.BytesIO()
    downloader = MediaIoBaseDownload(buf, request)
    done = False
    while not done:
        _, done = downloader.next_chunk()
    buf.seek(0)
    return buf.read()


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("file_id")
    ap.add_argument("--sheet", default=None, help="nome da aba especifica")
    ap.add_argument("--rows", type=int, default=10, help="qtd de linhas para imprimir por aba")
    ap.add_argument("--save", default=None, help="salvar xlsx baixado em arquivo local")
    args = ap.parse_args()

    blob = download_xlsx(args.file_id)
    if args.save:
        Path(args.save).write_bytes(blob)
        print(f"[ok] salvo em {args.save}", file=sys.stderr)

    wb = load_workbook(io.BytesIO(blob), data_only=True)
    print(f"=== Abas ({len(wb.sheetnames)}): {wb.sheetnames}")
    print()

    target_sheets = [args.sheet] if args.sheet else wb.sheetnames
    for name in target_sheets:
        if name not in wb.sheetnames:
            print(f"[warn] aba {name!r} nao existe", file=sys.stderr)
            continue
        ws = wb[name]
        print(f"--- Aba: {name} ({ws.max_row} linhas x {ws.max_column} colunas)")
        count = 0
        for row in ws.iter_rows(values_only=True):
            if count >= args.rows:
                print(f"    (... +{ws.max_row - args.rows} linhas)")
                break
            # formato compacto
            clean = [("" if v is None else str(v)[:60]) for v in row]
            print("  | ".join(clean))
            count += 1
        print()


if __name__ == "__main__":
    main()
