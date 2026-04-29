"""
parse_cronograma.py — Baixa o xlsx de cronograma do projeto e extrai
[cidade > escola > total alunos] para alimentar a seção Cronograma da landing.

Estratégia:
1. drive_find_cronograma → spreadsheet_id
2. Drive API export (.xlsx)
3. openpyxl: detecta abas-cidade pelo header e ignora as outras
4. Forward-fill em Data e Escola (merged cells)
5. Agrega Σ alunos por escola (descartando linhas-totalizador)
6. Salva JSON em output/cronogramas/{slug}.json

Usage:
  python tools/integrations/parse_cronograma.py --projeto 119
"""
import argparse
import io
import json
import re
import sys
import unicodedata
from datetime import date, datetime
from pathlib import Path

sys.path.insert(0, str(Path(__file__).parent))
from drive_find_cronograma import find_cronograma  # noqa: E402
from upload_to_drive import get_drive_service  # noqa: E402

import openpyxl
from googleapiclient.http import MediaIoBaseDownload

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")
    sys.stderr.reconfigure(encoding="utf-8", errors="replace")

REPO_ROOT = Path(__file__).resolve().parents[2]
TMP_DIR = REPO_ROOT / ".tmp"
OUTPUT_DIR = REPO_ROOT / "output" / "cronogramas"

# Abas que NUNCA são abas-cidade
SKIP_SHEETS = {
    "ENGAJAMENTO",
    "FORNECEDORES", "CHECK LIST DE COMPRAS", "CHECKLIST DE COMPRAS",
    "CHECK LIST", "CHECKLIST",
}


def _norm(s: str) -> str:
    if s is None:
        return ""
    s = str(s).strip()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s.upper()


def _download_xlsx(spreadsheet_id: str, mime_type: str, dest: Path) -> Path:
    service = get_drive_service()
    if mime_type == "application/vnd.google-apps.spreadsheet":
        # Google Sheet → exporta como xlsx
        req = service.files().export_media(
            fileId=spreadsheet_id,
            mimeType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        req = service.files().get_media(fileId=spreadsheet_id, supportsAllDrives=True)
    dest.parent.mkdir(parents=True, exist_ok=True)
    fh = io.FileIO(dest, "wb")
    dl = MediaIoBaseDownload(fh, req)
    done = False
    while not done:
        _, done = dl.next_chunk()
    return dest


def _is_cidade_sheet(ws) -> bool:
    """Heurística: aba-cidade tem header com Data + Escola + Atividade + Quantidade."""
    if _norm(ws.title) in {_norm(s) for s in SKIP_SHEETS}:
        return False
    header_tokens = []
    for row in ws.iter_rows(min_row=1, max_row=8, values_only=True):
        for cell in row:
            if cell is not None:
                header_tokens.append(_norm(cell))
    has_escola = any(("ESCOLA" in t or t == "LOCAL") for t in header_tokens)
    has_data = any(t == "DATA" or t.startswith("DATA ") for t in header_tokens)
    has_ativ = any("ATIVIDADE" in t for t in header_tokens)
    has_qtd = any("QUANTIDADE" in t or "ALUNOS" in t for t in header_tokens)
    return has_data and has_escola and has_ativ and has_qtd


def _detect_columns(ws) -> dict:
    """Retorna mapping {logical_name: col_index} a partir do header."""
    cols = {}
    # header pode estar nas primeiras 7 linhas (algumas planilhas têm título antes)
    for header_row in range(1, 8):
        for row in ws.iter_rows(min_row=header_row, max_row=header_row, values_only=False):
            for cell in row:
                v = _norm(cell.value)
                if not v:
                    continue
                if "DATA" in v and "data" not in cols:
                    cols["data"] = cell.column
                elif ("ESCOLA" in v or v == "LOCAL") and "escola" not in cols:
                    cols["escola"] = cell.column
                elif ("ATIVIDADE" in v or "ATIVIDADES" in v) and "atividade" not in cols:
                    cols["atividade"] = cell.column
                elif v == "CIDADE" and "cidade_col" not in cols:
                    cols["cidade_col"] = cell.column
                elif ("QUANTIDADE" in v or v.startswith("ALUNOS") or "QTD" in v) and "alunos" not in cols:
                    cols["alunos"] = cell.column
                elif "SERIE" in v and "serie" not in cols:
                    cols["serie"] = cell.column
                elif "PROJETO" in v and "CIDADE" in v and "cidade_inline" not in cols:
                    cols["cidade_inline"] = cell.column
        if {"data", "escola", "atividade", "alunos"} <= cols.keys():
            cols["header_row"] = header_row
            return cols
        cols.clear()
    return {}


def _parse_int(v) -> int | None:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        return int(v)
    s = str(v).strip().replace(".", "").replace(",", "")
    if not s.isdigit():
        return None
    return int(s)


def _parse_date(v):
    if v is None:
        return None
    if isinstance(v, (datetime, date)):
        return v.date() if isinstance(v, datetime) else v
    s = str(v).strip()
    for fmt in ("%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def parse_cidade_sheet(ws, cidade_default: str) -> list[dict]:
    """Extrai lista de turmas {data, escola, atividade, alunos, serie} da aba."""
    cols = _detect_columns(ws)
    if not cols:
        return []
    start_row = cols["header_row"] + 1
    out = []
    last_data, last_escola, last_cidade = None, None, None
    for row in ws.iter_rows(min_row=start_row, values_only=True):
        # converte 1-indexed → 0-indexed
        def cell(name):
            idx = cols.get(name)
            return row[idx - 1] if idx and idx - 1 < len(row) else None

        data = _parse_date(cell("data")) or last_data
        escola = str(cell("escola") or "").strip() or None
        if escola:
            last_escola = escola
        else:
            escola = last_escola
        # cidade pode vir da coluna ou do nome da aba (default)
        cidade_row = str(cell("cidade_col") or "").strip() or None
        if cidade_row:
            last_cidade = cidade_row
        cidade_eff = last_cidade or cidade_default
        atividade = str(cell("atividade") or "").strip()
        alunos = _parse_int(cell("alunos"))
        serie = str(cell("serie") or "").strip()

        if data:
            last_data = data
        # Linha-totalizador: tem alunos mas não tem atividade nem série nem escola novos
        # → descarta. Linha válida: tem atividade E (escola OU ainda forward-fill).
        if not atividade or not escola or alunos is None:
            continue
        # filtra escolas placeholder ainda não definidas
        if _norm(escola) in {"A DEFINIR", "DEFINIR", "A SER DEFINIDA"}:
            continue
        if not cidade_eff:
            continue  # sem cidade não dá pra agrupar
        out.append({
            "cidade": cidade_eff,
            "data": data.isoformat() if data else None,
            "escola": escola,
            "atividade": atividade,
            "serie": serie,
            "alunos": alunos,
        })
    return out


def _norm_escola(s: str) -> str:
    """Normaliza nome de escola: UPPER + colapsa espaços + tira acento."""
    s = re.sub(r"\s+", " ", str(s).strip()).upper()
    s = unicodedata.normalize("NFKD", s).encode("ascii", "ignore").decode("ascii")
    return s


def aggregate_by_school(turmas: list[dict]) -> list[dict]:
    """Agrupa cidade > escola > alunos únicos.

    Cada aluno costuma participar de várias atividades em dias diferentes,
    então NÃO somamos linha a linha. Em vez disso:
    - Para cada (cidade, escola, série), pegamos o MAX de alunos vistos.
    - Linhas sem série são tratadas como turma própria (cada linha conta uma vez).
    Total da escola = soma do max por turma.
    """
    # bucket por escola normalizada → guardar nome canônico (mais frequente)
    school_meta: dict[tuple[str, str], dict] = {}
    # subbucket por turma: {(cidade, escola_norm, serie): max_alunos}
    turma_max: dict[tuple[str, str, str], int] = {}
    # linhas-sem-série: cada uma soma uma vez, sem dedup
    sem_serie_total: dict[tuple[str, str], int] = {}

    for t in turmas:
        cidade = t["cidade"]
        escola_norm = _norm_escola(t["escola"])
        key_school = (cidade, escola_norm)
        meta = school_meta.setdefault(key_school, {
            "cidade": cidade, "escola_norm": escola_norm,
            "nomes": {}, "datas": set(), "atividades": set(),
        })
        # nome canônico = o mais comum (preserva caixa, mas colapsa espaços múltiplos)
        nome = re.sub(r"\s+", " ", t["escola"].strip())
        meta["nomes"][nome] = meta["nomes"].get(nome, 0) + 1
        if t["data"]:
            meta["datas"].add(t["data"])
        if t["atividade"]:
            meta["atividades"].add(t["atividade"])

        if t["serie"]:
            k = (cidade, escola_norm, _norm_escola(t["serie"]))
            turma_max[k] = max(turma_max.get(k, 0), t["alunos"])
        else:
            sem_serie_total[key_school] = sem_serie_total.get(key_school, 0) + t["alunos"]

    cards = []
    for key_school, meta in school_meta.items():
        total = sum(v for (c, e, _s), v in turma_max.items() if (c, e) == key_school)
        total += sem_serie_total.get(key_school, 0)
        # nome canônico = mais frequente
        nome_escola = max(meta["nomes"].items(), key=lambda x: x[1])[0]
        datas = sorted(meta["datas"])
        cards.append({
            "cidade": meta["cidade"],
            "escola": nome_escola,
            "total_alunos": total,
            "data_inicio": datas[0] if datas else None,
            "data_fim": datas[-1] if datas else None,
            "atividades": sorted(meta["atividades"]),
        })
    cards.sort(key=lambda c: (c["cidade"], c["escola"]))
    return cards


CIDADE_FALLBACK = {
    "125": "Guarulhos",  # GRU = todo projeto roda em Guarulhos
    "127G": "Guarulhos",
    "127S": "Guarulhos",
}


def parse_cronograma(projeto: str) -> dict:
    info = find_cronograma(projeto)
    if not info:
        raise SystemExit(f"projeto desconhecido: {projeto}")
    xlsx_path = TMP_DIR / f"cronograma_{projeto}.xlsx"
    print(f"Baixando {info['spreadsheet_name']} ...", file=sys.stderr)
    _download_xlsx(info["spreadsheet_id"], info["mime_type"], xlsx_path)

    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    all_turmas: list[dict] = []
    sheets_used = []
    for name in wb.sheetnames:
        ws = wb[name]
        if not _is_cidade_sheet(ws):
            continue
        # nome da aba pode vir como "Cronograma X", "X - GRADE EXECUÇÃO", "MOGI-GUAÇU"
        # nomes genéricos (sem cidade) → cidade_default = None; vai depender da coluna CIDADE
        clean = re.sub(r"[-_]", " ", name).strip()
        clean = re.sub(
            r"\b(?:cronograma|grade|execução|execucao|finalização|finalizacao|página\d*|pagina\d*|de\s+execução|de\s+execucao)\b",
            "", clean, flags=re.IGNORECASE).strip()
        # GRU sozinho = Guarulhos (sigla aeroporto/zona aéreo)
        if re.fullmatch(r"\s*GRU\s*", clean, re.IGNORECASE):
            clean = "Guarulhos"
        # se sobrou nada (ex: "GRADE DE EXECUÇÃO" → ""), usa fallback do projeto se houver
        cidade = clean.title() if clean else CIDADE_FALLBACK.get(projeto)
        turmas = parse_cidade_sheet(ws, cidade_default=cidade)
        if turmas:
            sheets_used.append({"sheet": name, "linhas": len(turmas)})
            all_turmas.extend(turmas)

    cards = aggregate_by_school(all_turmas)
    total_geral = sum(c["total_alunos"] for c in cards)

    return {
        "projeto": projeto,
        "slug": info["slug"],
        "fonte": {
            "spreadsheet_id": info["spreadsheet_id"],
            "spreadsheet_name": info["spreadsheet_name"],
            "last_modified": info["last_modified"],
        },
        "atualizado_em": date.today().isoformat(),
        "abas_processadas": sheets_used,
        "total_alunos": total_geral,
        "total_escolas": len(cards),
        "total_cidades": len({c["cidade"] for c in cards}),
        "cards": cards,
    }


def main():
    p = argparse.ArgumentParser()
    p.add_argument("--projeto", required=True)
    p.add_argument("--output", help="path JSON (default: output/cronogramas/{slug}.json)")
    args = p.parse_args()

    result = parse_cronograma(args.projeto)
    out = Path(args.output) if args.output else OUTPUT_DIR / f"{result['slug']}.json"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(result, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\nSalvo: {out}", file=sys.stderr)
    print(f"  {result['total_cidades']} cidade(s), {result['total_escolas']} escola(s), {result['total_alunos']} aluno(s)")
    for c in result["cards"]:
        print(f"  - {c['cidade']} / {c['escola']}: {c['total_alunos']} alunos")


if __name__ == "__main__":
    main()
