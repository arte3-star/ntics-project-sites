"""
Parser para planilha orçamentária exportada do SALIC.

Lê o arquivo Excel exportado do sistema SALIC (grid_PlanilhaOrcamentaria_Homologada.xlsx)
e extrai todas as rubricas em formato estruturado, prontas para uso no Engenhoca 2.0.

Uso:
    python tools/parse_salic_excel.py --input caminho/arquivo.xlsx [--output caminho/saida.json]

Saída: JSON com lista de rubricas no formato:
    {
        "fonte_recurso": "Incentivo Fiscal Federal",
        "produtos": [...],
        "rubricas": [
            {
                "numero": 1,
                "item": "AGENTE EDUCATIVO(A): PALESTRANTE",
                "produto": "OFICINA/WORKSHOP/SEMINÁRIO AUDIOVISUAL",
                "etapa": "2 - PRODUÇÃO / EXECUÇÃO",
                "uf": "SP",
                "cidade": "CAMPINAS",
                "nome_engenhoca": "AGENTE EDUCATIVO(A): PALESTRANTE - CAMPINAS/SP",
                "unidade": "DIA",
                "dias": 0,
                "quantidade": 1,
                "ocorrencia": 2,
                "valor_unitario": 1000.0,
                "valor_solicitado": 2000.0,
                "valor_sugerido": 2000.0,
                "valor_aprovado": 2000.0
            },
            ...
        ],
        "totais": {
            "total_geral": ...,
            "por_produto": {...},
            "por_etapa": {...}
        }
    }
"""

import argparse
import json
import re
import sys
from pathlib import Path

try:
    import openpyxl
except ImportError:
    print("Erro: openpyxl não instalado. Execute: pip install openpyxl")
    sys.exit(1)


def parse_context_line(cell_value: str) -> tuple[str, str] | None:
    """Detecta linhas de contexto hierárquico (ex: 'Produto : Exposição Cultural')."""
    if not cell_value or not isinstance(cell_value, str):
        return None

    text = cell_value.strip()

    # Padrões de contexto do SALIC
    patterns = [
        (r'^Fonte\s+Recurso\s*:\s*(.+)', 'fonte_recurso'),
        (r'^Produto\s*:\s*(.+)', 'produto'),
        (r'^Etapa\s*:\s*(.+)', 'etapa'),
        (r'^UF\s*:\s*(.+)', 'uf'),
        (r'^Municipio\s*:\s*(.+)', 'municipio'),
    ]

    for pattern, key in patterns:
        match = re.match(pattern, text, re.IGNORECASE)
        if match:
            return key, match.group(1).strip()

    return None


def is_header_row(row_values: list) -> bool:
    """Detecta linha de cabeçalho (Item | Unidade | Dias | ...)."""
    if not row_values or not row_values[0]:
        return False
    first = str(row_values[0]).strip()
    return first == 'Item'


def is_total_row(row_values: list) -> bool:
    """Detecta linhas de total (Total por município, etc.)."""
    if not row_values or not row_values[0]:
        return False
    first = str(row_values[0]).strip()
    return first.startswith('Total por')


def parse_salic_excel(file_path: str) -> dict:
    """
    Faz parse do arquivo Excel exportado do SALIC.

    Args:
        file_path: Caminho para o arquivo .xlsx

    Returns:
        Dicionário com rubricas extraídas e metadados
    """
    wb = openpyxl.load_workbook(file_path, data_only=True)

    # Procura a sheet de dados (geralmente "Consulta")
    sheet_name = None
    for name in wb.sheetnames:
        if name.lower() in ('consulta', 'sheet1', 'planilha1'):
            sheet_name = name
            break
    if not sheet_name:
        sheet_name = wb.sheetnames[0]

    ws = wb[sheet_name]

    # Estado do parser
    context = {
        'fonte_recurso': '',
        'produto': '',
        'etapa': '',
        'uf': '',
        'municipio': '',
    }

    rubricas = []
    produtos_set = set()
    etapas_set = set()
    numero = 0
    totais = {
        'por_produto': {},
        'por_etapa': {},
        'total_geral': 0,
    }

    for row in ws.iter_rows(values_only=True):
        # Converte row para lista de strings/valores
        values = list(row)

        # Pega o primeiro valor não-nulo
        first_val = values[0] if values else None

        if first_val is None:
            # Linha vazia, pula
            continue

        first_str = str(first_val).strip()

        if not first_str:
            continue

        # Tenta detectar linha de contexto
        ctx = parse_context_line(first_str)
        if ctx:
            key, value = ctx
            context[key] = value
            if key == 'produto':
                produtos_set.add(value)
            elif key == 'etapa':
                etapas_set.add(value)
            continue

        # Pula cabeçalhos
        if is_header_row(values):
            continue

        # Pula linhas de total
        if is_total_row(values):
            # Captura totais
            total_val = values[8] if len(values) > 8 and values[8] else 0
            if first_str.startswith('Total por produto'):
                totais['por_produto'][context['produto']] = float(total_val or 0)
            elif first_str.startswith('Total por etapa'):
                totais['por_etapa'][f"{context['produto']} | {context['etapa']}"] = float(total_val or 0)
            continue

        # Se chegou aqui, é uma linha de dados (rubrica)
        # Formato: Item | Unidade | Dias | Qtde | Ocorrência | Vl.Unitário | Vl.Solicitado | Vl.Sugerido | Vl.Aprovado
        item = first_str
        unidade = str(values[1]).strip() if len(values) > 1 and values[1] else ''
        dias = float(values[2] or 0) if len(values) > 2 and values[2] is not None else 0
        qtde = float(values[3] or 0) if len(values) > 3 and values[3] is not None else 0
        ocorrencia = float(values[4] or 0) if len(values) > 4 and values[4] is not None else 0
        vl_unitario = float(values[5] or 0) if len(values) > 5 and values[5] is not None else 0
        vl_solicitado = float(values[6] or 0) if len(values) > 6 and values[6] is not None else 0
        vl_sugerido = float(values[7] or 0) if len(values) > 7 and values[7] is not None else 0
        vl_aprovado = float(values[8] or 0) if len(values) > 8 and values[8] is not None else 0

        # Valida que parece ser um item real (tem unidade e valor)
        if not unidade and vl_aprovado == 0:
            continue

        numero += 1

        # Monta nome no formato Engenhoca: ITEM - CIDADE/UF
        cidade = context['municipio'].upper()
        uf = context['uf'].upper()
        nome_engenhoca = f"{item.upper()} - {cidade}/{uf}" if cidade else item.upper()

        rubrica = {
            'numero': numero,
            'item': item.upper(),
            'produto': context['produto'].upper(),
            'etapa': context['etapa'].upper(),
            'uf': uf,
            'cidade': cidade,
            'nome_engenhoca': nome_engenhoca,
            'unidade': unidade.upper(),
            'dias': dias,
            'quantidade': qtde,
            'ocorrencia': ocorrencia,
            'valor_unitario': vl_unitario,
            'valor_solicitado': vl_solicitado,
            'valor_sugerido': vl_sugerido,
            'valor_aprovado': vl_aprovado,
        }

        rubricas.append(rubrica)
        totais['total_geral'] += vl_aprovado

    # Extrai lista de produtos e etapas únicos (na ordem de aparição)
    produtos = sorted(produtos_set)
    etapas = sorted(etapas_set)

    result = {
        'fonte_recurso': context['fonte_recurso'],
        'produtos': produtos,
        'etapas': etapas,
        'total_rubricas': len(rubricas),
        'total_geral': totais['total_geral'],
        'totais': totais,
        'rubricas': rubricas,
    }

    return result


def print_summary(data: dict):
    """Imprime resumo dos dados extraídos."""
    print(f"\n{'='*60}")
    print(f"RESUMO DA PLANILHA SALIC")
    print(f"{'='*60}")
    print(f"Fonte: {data['fonte_recurso']}")
    print(f"Total de rubricas: {data['total_rubricas']}")
    print(f"Valor total aprovado: R$ {data['total_geral']:,.2f}")
    print(f"\nProdutos ({len(data['produtos'])}):")
    for p in data['produtos']:
        total = data['totais']['por_produto'].get(p, 0)
        print(f"  - {p} (R$ {total:,.2f})")
    print(f"\nPrimeiras 10 rubricas:")
    for r in data['rubricas'][:10]:
        print(f"  {r['numero']:3d}. {r['nome_engenhoca'][:60]:<60} | {r['unidade']:<10} | R$ {r['valor_aprovado']:>12,.2f}")
    if len(data['rubricas']) > 10:
        print(f"  ... e mais {len(data['rubricas']) - 10} rubricas")
    print(f"{'='*60}\n")


def main():
    parser = argparse.ArgumentParser(
        description='Parser para planilha orçamentária do SALIC'
    )
    parser.add_argument(
        '--input', '-i',
        required=True,
        help='Caminho para o arquivo Excel exportado do SALIC'
    )
    parser.add_argument(
        '--output', '-o',
        help='Caminho para salvar o JSON de saída (padrão: .tmp/salic_parsed.json)'
    )
    parser.add_argument(
        '--summary', '-s',
        action='store_true',
        default=True,
        help='Exibir resumo no terminal'
    )

    args = parser.parse_args()

    # Valida input
    input_path = Path(args.input)
    if not input_path.exists():
        print(f"Erro: arquivo não encontrado: {input_path}")
        sys.exit(1)

    # Define output
    if args.output:
        output_path = Path(args.output)
    else:
        output_path = Path('.tmp/salic_parsed.json')

    output_path.parent.mkdir(parents=True, exist_ok=True)

    # Parse
    print(f"Lendo arquivo: {input_path}")
    data = parse_salic_excel(str(input_path))

    # Summary
    if args.summary:
        print_summary(data)

    # Save JSON
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(data, f, ensure_ascii=False, indent=2)

    print(f"JSON salvo em: {output_path}")
    return data


if __name__ == '__main__':
    main()
